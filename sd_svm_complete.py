import pandas as pd
import os
import tkinter as tk
from tkinter import filedialog
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
import numpy as np
import time


# ==========================================
# ตั้งค่าฟอนต์ภาษาไทยสำหรับกราฟ
# ==========================================
def set_thai_font():
    if os.name == 'nt':  # Windows
        plt.rcParams['font.family'] = 'Tahoma'
    else:  # Mac / Linux
        plt.rcParams['font.family'] = 'Ayuthaya'
    plt.rcParams['axes.unicode_minus'] = False


# ==========================================
# ส่วนที่ 1: ฟังก์ชันเปลี่ยนชื่อไฟล์ (Rename)
# ==========================================
def batch_rename_mode():
    print("\n" + "=" * 60)
    print("   📂 โหมดเปลี่ยนชื่อไฟล์อัตโนมัติ (เรียงตามเวลาเก่า -> ใหม่)")
    print("=" * 60)

    root = tk.Tk();
    root.withdraw();
    root.attributes('-topmost', True)
    print("⏳ กรุณาเลือกไฟล์ CSV ที่ต้องการเปลี่ยนชื่อ (เลือกได้หลายไฟล์)...")

    file_paths = filedialog.askopenfilenames(
        title="เลือกไฟล์ CSV ที่ต้องการเปลี่ยนชื่อ",
        filetypes=[("CSV Files", "*.csv"), ("All Files", "*.*")]
    )

    if not file_paths:
        print("❌ ไม่ได้เลือกไฟล์ ยกเลิกการทำงาน")
        return

    files_with_time = []
    for f_path in file_paths:
        timestamp = os.path.getmtime(f_path)
        files_with_time.append((f_path, timestamp))

    files_with_time.sort(key=lambda x: x[1])

    print(f"\n📦 พบไฟล์ทั้งหมด {len(files_with_time)} ไฟล์")
    print("-" * 60)

    base_name = input("กรุณาตั้งชื่อกลุ่มไฟล์ใหม่ (เช่น Walking, Test): ").strip()
    if not base_name: base_name = "Data"

    print("-" * 60)
    print("กำลังดำเนินการเปลี่ยนชื่อ...\n")

    count = 0
    for i, (old_path, timestamp) in enumerate(files_with_time, 1):
        directory = os.path.dirname(old_path)
        extension = os.path.splitext(old_path)[1]
        new_filename = f"{base_name}_{i:02d}{extension}"
        new_path = os.path.join(directory, new_filename)
        time_str = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(timestamp))

        try:
            if old_path == new_path:
                print(f"   ⚠️ ข้าม: {os.path.basename(old_path)} (ชื่อตรงกันอยู่แล้ว)")
                continue
            os.rename(old_path, new_path)
            print(f"   ✔ [{time_str}] {os.path.basename(old_path)}  --->  {new_filename}")
            count += 1
        except Exception as e:
            print(f"   ❌ Error เปลี่ยนชื่อไฟล์ {os.path.basename(old_path)}: {e}")

    print("\n" + "=" * 60)
    print(f"🎉 เปลี่ยนชื่อเสร็จสิ้นจำนวน {count} ไฟล์!")
    print("=" * 60)
    input("\nกด Enter เพื่อกลับสู่เมนูหลัก...")


# ==========================================
# ส่วนที่ 2: ฟังก์ชันวิเคราะห์ข้อมูล (Analyze)
# ==========================================
def get_experiment_info():
    scenarios = [
        "กรณีเดินปกติ", "กรณีเดินปกติช้า", "กรณีเดินเอียงขวา",
        "กรณีเดินเอียงซ้าย", "กรณีเดินเอียงซ้ายขวา", "กรณีเดินหยุดชะงัก"
    ]
    print("\n" + "=" * 50)
    print("   กรุณาเลือกสถานการณ์การทดลอง (Scenario)")
    print("=" * 50)
    for index, scenario in enumerate(scenarios, 1):
        print(f" [{index}] {scenario}")
    print("-" * 50)

    while True:
        choice = input("เลือกหมายเลข (1-6): ").strip()
        if choice.isdigit() and 1 <= int(choice) <= 6:
            selected_scenario = scenarios[int(choice) - 1]
            break
        else:
            print("❌ ข้อมูลไม่ถูกต้อง")
    print("-" * 50)
    subject_name = input("กรุณาระบุชื่อผู้เข้าทดลอง: ").strip()
    return f"{selected_scenario} : {subject_name}"


def get_unique_filename(filepath):
    if not os.path.exists(filepath): return filepath
    base, ext = os.path.splitext(filepath)
    counter = 1
    while True:
        new_filepath = f"{base} ({counter}){ext}"
        if not os.path.exists(new_filepath): return new_filepath
        counter += 1


def parse_time_column(df):
    if 'Time' not in df.columns: return df.index.to_numpy()
    try:
        time_str = df['Time'].astype(str)

        def time_to_seconds(t_str):
            parts = t_str.strip().split(':')
            if len(parts) == 2:
                return float(parts[0]) * 60 + float(parts[1])
            elif len(parts) == 3:
                return float(parts[0]) * 3600 + float(parts[1]) * 60 + float(parts[2])
            else:
                return 0.0

        seconds_array = time_str.apply(time_to_seconds).to_numpy()
        if len(seconds_array) > 0:
            return seconds_array - seconds_array[0]
        else:
            return df.index.to_numpy()
    except Exception:
        return df.index.to_numpy()


def format_excel_table(worksheet, min_row, max_row, min_col, max_col):
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    for row in worksheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if isinstance(cell.value, (int, float)): cell.number_format = '0.0000'
    worksheet.column_dimensions['A'].width = 20
    for col_idx in range(min_col + 1, max_col + 1):
        worksheet.column_dimensions[get_column_letter(col_idx)].width = 15
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    for col_idx in range(min_col, max_col + 1):
        cell = worksheet.cell(row=min_row, column=col_idx)
        cell.font = Font(bold=True);
        cell.fill = header_fill


def calculate_svm(df, prefix):
    cols = [f'{prefix}_X', f'{prefix}_Y', f'{prefix}_Z']
    if all(col in df.columns for col in cols):
        df[f'{prefix}_SVM'] = (df[cols[0]] ** 2 + df[cols[1]] ** 2 + df[cols[2]] ** 2) ** 0.5
        return f'{prefix}_SVM'
    return None


def plot_svm_graph(df, filename_base, header_title, time_axis):
    try:
        set_thai_font()
        SMOOTH_WINDOW = 7
        fig, axs = plt.subplots(3, 1, figsize=(12, 12))
        fig.suptitle(f'ผลการวิเคราะห์ SVM : {header_title}', fontsize=16, fontweight='bold')

        # --- จุดที่แก้ไข: ใช้ MaxNLocator(integer=True) ---
        # เพื่อป้องกัน Error เมื่อไฟล์ยาวเกินไป (เกิน 1000 วินาที)
        # ตัวเลือกนี้จะบังคับให้เป็นจำนวนเต็มเสมอ แต่จะปรับระยะห่างให้อัตโนมัติ
        locator = ticker.MaxNLocator(integer=True)

        def smooth_data(series, window):
            return series.rolling(window=window, center=True, min_periods=1).mean()

        # Plot ACC
        if 'ACC_SVM' in df.columns:
            smoothed = smooth_data(df['ACC_SVM'], SMOOTH_WINDOW)
            axs[0].plot(time_axis, smoothed, color='blue', label=f'ACC SVM (Smooth={SMOOTH_WINDOW})', linewidth=1.5)
            axs[0].set_ylabel('Acceleration (g)')
            axs[0].set_title('Accelerometer SVM (ขนาดแรงรวมความเร่ง - แบบมน)', fontsize=12)
            axs[0].set_xlabel('Time (seconds)')
            axs[0].xaxis.set_major_locator(locator)  # ใช้ Locator แบบอัตโนมัติ
            axs[0].grid(True, linestyle='--', alpha=0.6);
            axs[0].legend(loc='upper right')

        # Plot GYRO
        if 'GYRO_SVM' in df.columns:
            smoothed = smooth_data(df['GYRO_SVM'], SMOOTH_WINDOW)
            axs[1].plot(time_axis, smoothed, color='green', label=f'GYRO SVM (Smooth={SMOOTH_WINDOW})', linewidth=1.5)
            axs[1].set_ylabel('Angular Velocity (deg/s)')
            axs[1].set_title('Gyroscope SVM (ขนาดแรงรวมไจโร - แบบมน)', fontsize=12)
            axs[1].set_xlabel('Time (seconds)')
            axs[1].xaxis.set_major_locator(locator)  # ใช้ Locator แบบอัตโนมัติ
            axs[1].grid(True, linestyle='--', alpha=0.6);
            axs[1].legend(loc='upper right')

        # Plot MAG
        if 'MAG_SVM' in df.columns:
            smoothed = smooth_data(df['MAG_SVM'], SMOOTH_WINDOW)
            axs[2].plot(time_axis, smoothed, color='red', label=f'MAG SVM (Smooth={SMOOTH_WINDOW})', linewidth=1.5)
            axs[2].set_ylabel('Magnetic Field (uT)')
            axs[2].set_title('Magnetometer SVM (ขนาดแรงรวมแม่เหล็ก - แบบมน)', fontsize=12)
            axs[2].set_xlabel('Time (seconds)')
            axs[2].xaxis.set_major_locator(locator)  # ใช้ Locator แบบอัตโนมัติ
            axs[2].grid(True, linestyle='--', alpha=0.6);
            axs[2].legend(loc='upper right')

        plt.tight_layout(rect=[0, 0.03, 1, 0.96], h_pad=3.0)
        image_filename = filename_base + '_svm_plot_smooth.png'
        plt.savefig(image_filename)
        plt.close()
        return image_filename
    except Exception as e:
        print(f"   ⚠️ ไม่สามารถสร้างกราฟได้: {e}")
        return None


def analyze_data_mode():
    experiment_title = get_experiment_info()
    root = tk.Tk();
    root.withdraw();
    root.attributes('-topmost', True)
    print("\n⏳ กรุณาเลือกไฟล์ CSV ที่ต้องการวิเคราะห์ (เลือกได้หลายไฟล์)...")
    file_paths = filedialog.askopenfilenames(
        title="เลือกไฟล์ CSV ข้อมูลเซนเซอร์",
        filetypes=[("CSV Files", "*.csv"), ("All Files", "*.*")]
    )

    if file_paths:
        total_files = len(file_paths)
        print(f"\n📦 เลือกทั้งหมด: {total_files} ไฟล์")
        for i, file_path in enumerate(file_paths, 1):
            process_sensor_file(file_path, experiment_title, i, total_files)
        print("\n" + "=" * 50)
        print("🎉 ประมวลผลเสร็จสมบูรณ์!")
        print("=" * 50)
        input("\nกด Enter เพื่อกลับสู่เมนูหลัก...")
    else:
        print("❌ ไม่ได้เลือกไฟล์")


def process_sensor_file(file_path, header_title, current_idx, total_files):
    filename = os.path.basename(file_path)
    print(f"\n[{current_idx}/{total_files}] 📄 กำลังประมวลผล: {filename}")
    try:
        df = pd.read_csv(file_path)
        elapsed_time = parse_time_column(df)
        svm_cols = []
        for sensor in ['ACC', 'GYRO', 'MAG']:
            new_col = calculate_svm(df, sensor)
            if new_col: svm_cols.append(new_col)

        target_columns = [
            'ACC_X', 'ACC_Y', 'ACC_Z', 'ACC_SVM',
            'GYRO_X', 'GYRO_Y', 'GYRO_Z', 'GYRO_SVM',
            'MAG_X', 'MAG_Y', 'MAG_Z', 'MAG_SVM'
        ]
        valid_columns = [col for col in target_columns if col in df.columns]
        if not valid_columns:
            print(f"   ❌ ข้ามไฟล์นี้: ไม่พบคอลัมน์ข้อมูลที่ต้องการ")
            return

        stats_data = {}
        for col in valid_columns:
            series = pd.to_numeric(df[col], errors='coerce')
            stats_data[col] = {
                'Mean': series.mean(), 'Median': series.median(), 'SD': series.std(),
                'Skewness': series.skew(), 'Kurtosis': series.kurt(),
                'Max': series.max(), 'Min': series.min()
            }
        summary_df = pd.DataFrame(stats_data)

        base_filename_full = os.path.splitext(file_path)[0] + '_summary.xlsx'
        final_excel_filename = get_unique_filename(base_filename_full)
        final_base_name = os.path.splitext(final_excel_filename)[0]

        start_row = 3
        with pd.ExcelWriter(final_excel_filename, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Raw Data', index=False)
            summary_df.to_excel(writer, sheet_name='Summary Stats', startrow=start_row - 1)
            workbook = writer.book;
            worksheet = writer.sheets['Summary Stats']
            worksheet['A1'] = "การทดลอง:";
            worksheet['B1'] = header_title
            worksheet['A1'].font = Font(bold=True, size=12);
            worksheet['B1'].font = Font(bold=True, size=12)
            worksheet['D1'] = "Time Source: CSV File (Real-time)";
            worksheet['D1'].font = Font(italic=True, color="555555")
            max_r = start_row + summary_df.shape[0];
            max_c = summary_df.shape[1] + 1
            format_excel_table(worksheet, min_row=start_row, max_row=max_r, min_col=1, max_col=max_c)
        print(f"   ✔ บันทึก Excel: {os.path.basename(final_excel_filename)}")

        image_file = plot_svm_graph(df, final_base_name, header_title, elapsed_time)
        if image_file: print(f"   ✔ บันทึกกราฟ: {os.path.basename(image_file)}")

    except Exception as e:
        print(f"   ❌ เกิดข้อผิดพลาด: {e}")


if __name__ == "__main__":
    while True:
        print("\n" + "#" * 50)
        print("   🤖 โปรแกรมวิเคราะห์ข้อมูลเซนเซอร์ (Sensor Analysis)")
        print("#" * 50)
        print(" [1] 🏷️  เปลี่ยนชื่อไฟล์ (เรียงตามเวลาที่สร้าง)")
        print(" [2] 📊 วิเคราะห์ข้อมูล (คำนวณ SVM + กราฟ + Excel)")
        print(" [0] ❌ ออกจากโปรแกรม")
        print("-" * 50)

        choice = input("👉 เลือกเมนู (0-2): ").strip()
        if choice == '1':
            batch_rename_mode()
        elif choice == '2':
            analyze_data_mode()
        elif choice == '0':
            print("👋 บ๊ายบาย!");
            break
        else:
            print("❌ กรุณาเลือกเลข 0, 1 หรือ 2 เท่านั้น")