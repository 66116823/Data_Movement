import pandas as pd
import os
import tkinter as tk
from tkinter import filedialog
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
import numpy as np
import time


# ==========================================
# ตั้งค่าฟอนต์ภาษาไทยสำหรับกราฟ
# ==========================================
def set_thai_font():
    if os.name == 'nt':  # Windows
        plt.rcParams['font.family'] = 'Tahoma'
    else:  # Mac / Linux
        plt.rcParams['font.family'] = 'Ayuthaya'
    plt.rcParams['axes.unicode_minus'] = False


# ==========================================
# ส่วนที่ 1: ฟังก์ชันคำนวณและ Normalization
# ==========================================

def calculate_svm(df, prefix):
    """คำนวณ SVM (Raw Magnitude)"""
    cols = [f'{prefix}_X', f'{prefix}_Y', f'{prefix}_Z']
    if all(col in df.columns for col in cols):
        svm_col = f'{prefix}_SVM'
        # สูตร: sqrt(x^2 + y^2 + z^2)
        df[svm_col] = (df[cols[0]] ** 2 + df[cols[1]] ** 2 + df[cols[2]] ** 2) ** 0.5
        return svm_col
    return None


def normalize_series(series):
    """
    Min-Max Normalization: ปรับค่าให้อยู่ในช่วง 0 ถึง 1
    สูตร: (X - Min) / (Max - Min)
    """
    if series.max() == series.min():  # ป้องกันการหารด้วย 0 กรณีค่าเท่ากันหมด
        return series.apply(lambda x: 0.0)
    return (series - series.min()) / (series.max() - series.min())


def parse_time_column(df):
    """แปลงเวลาในไฟล์เป็นวินาทีเริ่มที่ 0"""
    if 'Time' not in df.columns: return df.index.to_numpy()
    try:
        time_str = df['Time'].astype(str)

        def time_to_seconds(t_str):
            parts = t_str.strip().split(':')
            if len(parts) == 2:
                return float(parts[0]) * 60 + float(parts[1])
            elif len(parts) == 3:
                return float(parts[0]) * 3600 + float(parts[1]) * 60 + float(parts[2])
            else:
                return 0.0

        seconds_array = time_str.apply(time_to_seconds).to_numpy()
        if len(seconds_array) > 0:
            return seconds_array - seconds_array[0]
        else:
            return df.index.to_numpy()
    except Exception:
        return df.index.to_numpy()


# ==========================================
# ส่วนที่ 2: ฟังก์ชันจัดการไฟล์และ Excel
# ==========================================

def get_unique_filename(filepath):
    if not os.path.exists(filepath): return filepath
    base, ext = os.path.splitext(filepath)
    counter = 1
    while True:
        new_filepath = f"{base} ({counter}){ext}"
        if not os.path.exists(new_filepath): return new_filepath
        counter += 1


def format_excel_table(worksheet, min_row, max_row, min_col, max_col):
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    for row in worksheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if isinstance(cell.value, (int, float)): cell.number_format = '0.0000'
    worksheet.column_dimensions['A'].width = 20
    for col_idx in range(min_col + 1, max_col + 1):
        worksheet.column_dimensions[get_column_letter(col_idx)].width = 16
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    for col_idx in range(min_col, max_col + 1):
        cell = worksheet.cell(row=min_row, column=col_idx)
        cell.font = Font(bold=True);
        cell.fill = header_fill


# ==========================================
# ส่วนที่ 3: ฟังก์ชัน Plot กราฟ (Normalized)
# ==========================================

def plot_normalized_graph(df, filename_base, header_title, time_axis):
    try:
        set_thai_font()
        SMOOTH_WINDOW = 7  # ค่าความมนของกราฟ

        fig, axs = plt.subplots(3, 1, figsize=(12, 12))
        # Title ระบุชัดเจนว่าเป็น Normalized
        fig.suptitle(f'Normalized SVM Analysis (0-1 Scale): {header_title}', fontsize=16, fontweight='bold')

        locator = ticker.MaxNLocator(integer=True)  # ป้องกัน Error ไฟล์ยาว

        def smooth_data(series, window):
            return series.rolling(window=window, center=True, min_periods=1).mean()

        # Helper function ในการวาดแต่ละแกน
        def plot_subplot(ax, data_col, color, label_text, title_text):
            if data_col in df.columns:
                # 1. ใช้ข้อมูล Normalized
                data_to_plot = df[data_col]
                # 2. ทำ Smoothing
                smoothed = smooth_data(data_to_plot, SMOOTH_WINDOW)

                ax.plot(time_axis, smoothed, color=color, label=f'Norm SVM (Smooth={SMOOTH_WINDOW})', linewidth=1.5)
                ax.set_ylabel('Norm. Mag (0-1)')  # เปลี่ยนหน่วยเป็น 0-1
                ax.set_title(title_text, fontsize=12)
                ax.set_xlabel('Time (seconds)')
                ax.set_ylim(-0.05, 1.05)  # ล็อคแกน Y ให้เห็น 0 ถึง 1 ชัดๆ
                ax.xaxis.set_major_locator(locator)
                ax.grid(True, linestyle='--', alpha=0.6)
                ax.legend(loc='upper right')

        # Plot ACC (Normalized)
        plot_subplot(axs[0], 'ACC_SVM_Norm', 'blue', 'ACC', 'Accelerometer: Normalized Pattern (รูปแบบความเร่ง)')

        # Plot GYRO (Normalized)
        plot_subplot(axs[1], 'GYRO_SVM_Norm', 'green', 'GYRO', 'Gyroscope: Normalized Pattern (รูปแบบการหมุน)')

        # Plot MAG (Normalized)
        plot_subplot(axs[2], 'MAG_SVM_Norm', 'red', 'MAG', 'Magnetometer: Normalized Pattern (รูปแบบแม่เหล็ก)')

        plt.tight_layout(rect=[0, 0.03, 1, 0.96], h_pad=3.0)

        # บันทึกเป็นชื่อ _norm_plot เพื่อแยกกับกราฟเก่า
        image_filename = filename_base + '_norm_plot.png'
        plt.savefig(image_filename)
        plt.close()
        return image_filename
    except Exception as e:
        print(f"   ⚠️ ไม่สามารถสร้างกราฟได้: {e}")
        return None


# ==========================================
# ส่วนที่ 4: Logic หลัก (Rename + Analyze)
# ==========================================

def batch_rename_mode():
    print("\n" + "=" * 60);
    print("   📂 โหมดเปลี่ยนชื่อไฟล์อัตโนมัติ");
    print("=" * 60)
    root = tk.Tk();
    root.withdraw();
    root.attributes('-topmost', True)
    file_paths = filedialog.askopenfilenames(title="เลือกไฟล์ CSV", filetypes=[("CSV Files", "*.csv")])
    if not file_paths: return

    files_with_time = [(f, os.path.getmtime(f)) for f in file_paths]
    files_with_time.sort(key=lambda x: x[1])

    base_name = input("ตั้งชื่อกลุ่มไฟล์ใหม่ (เช่น Patient_A): ").strip() or "Data"
    count = 0
    for i, (old_path, timestamp) in enumerate(files_with_time, 1):
        directory = os.path.dirname(old_path)
        new_filename = f"{base_name}_{i:02d}{os.path.splitext(old_path)[1]}"
        new_path = os.path.join(directory, new_filename)
        try:
            if old_path != new_path:
                os.rename(old_path, new_path)
                print(f"   ✔ {os.path.basename(old_path)} -> {new_filename}")
                count += 1
        except Exception as e:
            print(f"Error: {e}")
    print(f"🎉 เสร็จสิ้น {count} ไฟล์!");
    input("กด Enter กลับเมนู...")


def get_experiment_info():
    scenarios = ["กรณีเดินปกติ", "กรณีเดินปกติช้า", "กรณีเดินเอียงขวา", "กรณีเดินเอียงซ้าย", "กรณีเดินเอียงซ้ายขวา",
                 "กรณีเดินหยุดชะงัก"]
    print("\n--- เลือกสถานการณ์ ---")
    for i, s in enumerate(scenarios, 1): print(f" [{i}] {s}")
    while True:
        c = input("เลือก (1-6): ").strip()
        if c.isdigit() and 1 <= int(c) <= 6: return f"{scenarios[int(c) - 1]} : {input('ชื่อผู้ทดลอง: ').strip()}"


def process_sensor_file(file_path, header_title, current_idx, total_files):
    filename = os.path.basename(file_path)
    print(f"\n[{current_idx}/{total_files}] 📄 กำลังประมวลผล: {filename}")
    try:
        df = pd.read_csv(file_path)
        elapsed_time = parse_time_column(df)

        # 1. คำนวณ Raw SVM และ 2. สร้าง Normalized SVM Column
        svm_cols = []
        for sensor in ['ACC', 'GYRO', 'MAG']:
            raw_col = calculate_svm(df, sensor)  # ได้ ACC_SVM
            if raw_col:
                svm_cols.append(raw_col)
                # สร้างคอลัมน์ใหม่ เช่น ACC_SVM_Norm
                norm_col = f'{raw_col}_Norm'
                df[norm_col] = normalize_series(df[raw_col])

        # เลือกคอลัมน์ที่จะคำนวณสถิติ (เอาเฉพาะค่าจริง ไม่เอา Norm เพราะดูยากในตาราง)
        target_stats_cols = [
            'ACC_X', 'ACC_Y', 'ACC_Z', 'ACC_SVM',
            'GYRO_X', 'GYRO_Y', 'GYRO_Z', 'GYRO_SVM',
            'MAG_X', 'MAG_Y', 'MAG_Z', 'MAG_SVM'
        ]
        valid_stats_cols = [c for c in target_stats_cols if c in df.columns]

        if not valid_stats_cols:
            print("   ❌ ไม่พบคอลัมน์ข้อมูล")
            return

        # คำนวณสถิติ (Mean, Max, Min ของค่าจริง หน่วย g/deg)
        stats_data = {}
        for col in valid_stats_cols:
            series = pd.to_numeric(df[col], errors='coerce')
            stats_data[col] = {
                'Mean': series.mean(), 'Median': series.median(), 'SD': series.std(),
                'Skewness': series.skew(), 'Kurtosis': series.kurt(),
                'Max': series.max(), 'Min': series.min()
            }
        summary_df = pd.DataFrame(stats_data)

        # บันทึก Excel
        base_filename_full = os.path.splitext(file_path)[0] + '_summary.xlsx'
        final_excel_filename = get_unique_filename(base_filename_full)
        final_base_name = os.path.splitext(final_excel_filename)[0]

        start_row = 3
        with pd.ExcelWriter(final_excel_filename, engine='openpyxl') as writer:
            # Save Raw Data รวมคอลัมน์ Norm ไปด้วยเผื่ออยากดู
            df.to_excel(writer, sheet_name='Raw Data', index=False)
            summary_df.to_excel(writer, sheet_name='Summary Stats', startrow=start_row - 1)

            ws = writer.sheets['Summary Stats']
            ws['A1'] = "การทดลอง:";
            ws['B1'] = header_title
            ws['A1'].font = Font(bold=True, size=12);
            ws['B1'].font = Font(bold=True, size=12)
            ws['D1'] = "Graph Type: Normalized (0-1)";
            ws['D1'].font = Font(italic=True, color="555555")

            max_r = start_row + summary_df.shape[0];
            max_c = summary_df.shape[1] + 1
            format_excel_table(ws, min_row=start_row, max_row=max_r, min_col=1, max_col=max_c)
        print(f"   ✔ Excel (Stats from Raw Data): {os.path.basename(final_excel_filename)}")

        # วาดกราฟ (ใช้ข้อมูล Normalized)
        image_file = plot_normalized_graph(df, final_base_name, header_title, elapsed_time)
        if image_file: print(f"   ✔ Graph (Normalized 0-1): {os.path.basename(image_file)}")

    except Exception as e:
        print(f"   ❌ Error: {e}")


def analyze_data_mode():
    title = get_experiment_info()
    root = tk.Tk();
    root.withdraw();
    root.attributes('-topmost', True)
    paths = filedialog.askopenfilenames(title="เลือกไฟล์ CSV", filetypes=[("CSV Files", "*.csv")])
    if paths:
        print(f"\n📦 เลือก {len(paths)} ไฟล์")
        for i, p in enumerate(paths, 1): process_sensor_file(p, title, i, len(paths))
        print("\n🎉 เสร็จสมบูรณ์!");
        input("กด Enter กลับเมนู...")


if __name__ == "__main__":
    while True:
        print("\n=== Sensor Analysis (Normalized SVM) ===")
        print(" [1] เปลี่ยนชื่อไฟล์ (Batch Rename)")
        print(" [2] วิเคราะห์ข้อมูล (Normalize + Plot)")
        print(" [0] ออก")
        c = input("เลือก: ").strip()
        if c == '1':
            batch_rename_mode()
        elif c == '2':
            analyze_data_mode()
        elif c == '0':
            break