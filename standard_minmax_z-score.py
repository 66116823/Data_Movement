import pandas as pd
import os
import shutil
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import matplotlib.pyplot as plt
import json


# ==========================================
# 1. ตั้งค่าพื้นฐาน
# ==========================================
def set_thai_font():
    if os.name == 'nt':
        plt.rcParams['font.family'] = 'Tahoma'
    else:
        plt.rcParams['font.family'] = 'Ayuthaya'
    plt.rcParams['axes.unicode_minus'] = False


CONFIG_FILE = 'global_standard_config.json'
GLOBAL_STATS = None


# ==========================================
# 2. Core Logic (Calculation)
# ==========================================
def parse_time_column(df):
    if 'Time' not in df.columns: return df.index.to_numpy()
    try:
        time_str = df['Time'].astype(str)

        def t2s(t):
            p = t.strip().split(':')
            if len(p) == 2:
                return float(p[0]) * 60 + float(p[1])
            elif len(p) == 3:
                return float(p[0]) * 3600 + float(p[1]) * 60 + float(p[2])
            return 0.0

        sec = time_str.apply(t2s).to_numpy()
        return sec - sec[0] if len(sec) > 0 else df.index.to_numpy()
    except:
        return df.index.to_numpy()


def calculate_features(df):
    t_sec = parse_time_column(df)
    features_df = pd.DataFrame()
    for s in ['ACC', 'GYRO', 'MAG']:
        cols = [f'{s}_X', f'{s}_Y', f'{s}_Z']
        if all(c in df.columns for c in cols):
            svm = (df[cols[0]] ** 2 + df[cols[1]] ** 2 + df[cols[2]] ** 2) ** 0.5
            features_df[f'{s}_SVM'] = svm
            dt_series = pd.Series(t_sec).diff()
            avg_dt = dt_series.mean() if dt_series.mean() > 0 else 0.02
            svm_smooth = svm.rolling(window=5, center=True).mean().bfill().ffill()
            jerk = (svm_smooth.diff() / avg_dt).abs().fillna(0)
            features_df[f'{s}_SVM_Jerk'] = jerk
            features_df[f'{s}_SVM_Hybrid'] = svm * jerk
    return features_df


# ==========================================
# 3. Global Standard Logic
# ==========================================
def build_global_standard():
    print("\n" + "=" * 60)
    print("   🏗️  BUILDING GLOBAL STANDARD")
    print("=" * 60)
    root = tk.Tk();
    root.withdraw();
    root.attributes('-topmost', True)
    folder_path = filedialog.askdirectory(title="เลือกโฟลเดอร์ข้อมูลทั้งหมด")
    if not folder_path: return

    stats_acc = {}
    files = []
    for r, d, f in os.walk(folder_path):
        for file in f:
            if file.lower().endswith('.csv'): files.append(os.path.join(r, file))

    print(f"\nพบไฟล์ทั้งหมด: {len(files)} ไฟล์ ... กำลังประมวลผล")
    for idx, fpath in enumerate(files, 1):
        try:
            print(f"[{idx}/{len(files)}] Reading: {os.path.basename(fpath)}", end='\r')
            df = pd.read_csv(fpath)
            feat_df = calculate_features(df)
            for col in feat_df.columns:
                data = feat_df[col].dropna()
                if data.empty: continue
                if col not in stats_acc:
                    stats_acc[col] = {'n': 0, 'sum': 0.0, 'sum_sq': 0.0, 'min': float('inf'), 'max': float('-inf')}
                s = stats_acc[col]
                s['n'] += len(data);
                s['sum'] += data.sum();
                s['sum_sq'] += (data ** 2).sum()
                s['min'] = min(s['min'], data.min());
                s['max'] = max(s['max'], data.max())
        except:
            pass

    final_config = {}
    for col, s in stats_acc.items():
        if s['n'] > 0:
            mean = s['sum'] / s['n']
            variance = (s['sum_sq'] / s['n']) - (mean ** 2)
            std = variance ** 0.5 if variance > 0 else 0
            final_config[col] = {'mean': mean, 'std': std, 'min': s['min'], 'max': s['max']}

    with open(CONFIG_FILE, 'w') as f:
        json.dump(final_config, f, indent=4)
    print(f"\n✅ บันทึกค่ามาตรฐานกลางเรียบร้อยแล้วที่: {CONFIG_FILE}")


def load_global_standard():
    global GLOBAL_STATS
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, 'r') as f:
            GLOBAL_STATS = json.load(f)
        print(f"\nℹ️ Loaded Global Standard (พร้อมใช้งาน)")
    else:
        GLOBAL_STATS = None
        print("\n⚠️ ไม่พบไฟล์มาตรฐานกลาง (จะใช้ Local Scaling)")


# ==========================================
# 4. Scaling & Plotting
# ==========================================
def scale_minmax(series, col_name):
    if GLOBAL_STATS and col_name in GLOBAL_STATS:
        g_min = GLOBAL_STATS[col_name]['min'];
        g_max = GLOBAL_STATS[col_name]['max']
        denom = g_max - g_min if g_max != g_min else 1.0
        return (series - g_min) / denom
    else:
        if series.max() == series.min(): return series.apply(lambda x: 0.0)
        return (series - series.min()) / (series.max() - series.min())


def scale_zscore(series, col_name):
    if GLOBAL_STATS and col_name in GLOBAL_STATS:
        g_mean = GLOBAL_STATS[col_name]['mean'];
        g_std = GLOBAL_STATS[col_name]['std']
        return (series - g_mean) / g_std if g_std != 0 else series.apply(lambda x: 0.0)
    else:
        std = series.std()
        return (series - series.mean()) / std if std != 0 else series.apply(lambda x: 0.0)


def plot_metric_comparison(df, filename_base, header_title, time_axis, metric_type):
    try:
        set_thai_font()
        suffix_map = {'SVM': '_SVM', 'Jerk': '_SVM_Jerk', 'Hybrid': '_SVM_Hybrid'}
        suffix = suffix_map.get(metric_type)
        if not suffix: return None

        fig, axs = plt.subplots(3, 2, figsize=(16, 12))
        mode_text = "Global Standard" if GLOBAL_STATS else "Local Scaling"
        fig.suptitle(f'{metric_type}: {header_title}\nMode: {mode_text}', fontsize=16, fontweight='bold')

        axs[0, 0].set_title(f"Min-Max (0-1)", fontsize=14, color='blue')
        axs[0, 1].set_title(f"Z-Score (SD)", fontsize=14, color='green')
        sensors = ['ACC', 'GYRO', 'MAG']

        for i, sensor in enumerate(sensors):
            col_name = f'{sensor}{suffix}'
            if col_name in df.columns:
                data = df[col_name]
                mm = scale_minmax(data, col_name)
                zs = scale_zscore(data, col_name)
                # Plot
                axs[i, 0].plot(time_axis, mm.rolling(7, center=True, min_periods=1).mean(), color='tab:blue')
                axs[i, 1].plot(time_axis, zs.rolling(7, center=True, min_periods=1).mean(), color='tab:green')

                axs[i, 0].set_ylabel(f'{sensor}');
                axs[i, 0].grid(True, alpha=0.3)
                axs[i, 1].grid(True, alpha=0.3);
                axs[i, 1].axhline(0, color='black', lw=0.5)
                if not GLOBAL_STATS: axs[i, 0].set_ylim(-0.05, 1.05)

        plt.tight_layout(rect=[0, 0.03, 1, 0.93])
        img_name = f"{filename_base}_{metric_type}.png"
        plt.savefig(img_name)
        plt.close()
        return img_name
    except:
        return None


# ==========================================
# 5. Helper: Auto Parsing Logic
# ==========================================
def parse_filename_info(filename):
    """
    แกะชื่อไฟล์ตามรูปแบบ: Scenario_Scenario_Name_ID.csv
    เช่น 'Dunk_Walk_ALIF_02.csv'
    -> Scenario: 'Dunk_Walk'
    -> Name: 'ALIF' (ใช้จัดกลุ่ม)
    -> ID: '02'
    """
    base = os.path.splitext(filename)[0]
    parts = base.split('_')

    # กรณีชื่อไฟล์สั้นเกินไป (ไม่มี _)
    if len(parts) < 2:
        return "Unknown_Scenario", "Unknown_Person", base

    # Logic:
    # ตัวสุดท้ายคือ ID
    # ตัวรองสุดท้ายคือ ชื่อคน (Key หลัก)
    # ที่เหลือข้างหน้าคือ Scenario
    try:
        id_str = parts[-1]
        name_str = parts[-2]
        scenario_str = "_".join(parts[:-2])

        # ถ้าไม่มี scenario (เช่น ALIF_01.csv) ให้ใช้ชื่อไฟล์เป็น scenario
        if not scenario_str: scenario_str = "General"

        return scenario_str, name_str, id_str
    except:
        return "Error", "Error", base


def format_excel(ws, min_r, max_r, min_c, max_c):
    border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    fill = PatternFill('solid', fgColor="D9D9D9")
    for r in ws.iter_rows(min_r, max_r, min_c, max_c):
        for c in r:
            c.border = border
            c.alignment = Alignment('center', 'center')
            if isinstance(c.value, float): c.number_format = '0.0000'
    for i in range(min_c, max_c + 1):
        c = ws.cell(min_r, i);
        c.font = Font(bold=True);
        c.fill = fill


# ==========================================
# 6. Main Process Logic (Auto Grouping)
# ==========================================
def process_single_file(path, scenario, name, file_id):
    try:
        df = pd.read_csv(path)
        feat_df = calculate_features(df)
        df = pd.concat([df, feat_df], axis=1)
        t_sec = parse_time_column(df)

        full_title = f"{scenario} : {name} (Trial {file_id})"
        xls_name = f"{scenario}_{name}_{file_id}_Report.xlsx"

        # Save Excel
        stats = {}
        for c in feat_df.columns:
            s = df[c]
            stats[c] = {'Mean': s.mean(), 'Max': s.max(), 'Min': s.min(), 'SD': s.std()}

        with pd.ExcelWriter(xls_name, engine='openpyxl') as w:
            df.to_excel(w, sheet_name='Data', index=False)
            sdf = pd.DataFrame(stats)
            sdf.to_excel(w, sheet_name='Stats', startrow=2)
            ws = w.sheets['Stats']
            ws['A1'] = "Title:";
            ws['B1'] = full_title
            if GLOBAL_STATS: ws['A2'] = "** GLOBAL SCALING **"
            format_excel(ws, 3, 3 + len(stats), 1, sdf.shape[1] + 1)

        # Plot
        files = [xls_name]
        base = os.path.splitext(xls_name)[0]
        for m in ['SVM', 'Jerk', 'Hybrid']:
            img = plot_metric_comparison(df, base, full_title, t_sec, m)
            if img: files.append(img)

        return files
    except Exception as e:
        print(f"Error {os.path.basename(path)}: {e}")
        return []


def analyze_auto_mode():
    load_global_standard()

    print("\n👉 กรุณาเลือกไฟล์ข้อมูล **ทั้งหมด** ที่ต้องการวิเคราะห์ (เลือกครอบจักรวาลได้เลย)")
    root = tk.Tk();
    root.withdraw();
    root.attributes('-topmost', True)
    file_paths = filedialog.askopenfilenames(title="เลือกไฟล์ CSV ทั้งหมด (ทุกคน)", filetypes=[("CSV", "*.csv")])

    if not file_paths: return

    # --- Step 1: Group Files by Person ---
    print("\n🔍 กำลังแยกกลุ่มข้อมูลจากชื่อไฟล์...")
    grouped_data = {}  # { 'ALIF': [path1, path2], 'Dunk': [path1...] }

    for path in file_paths:
        filename = os.path.basename(path)
        scenario, person_name, f_id = parse_filename_info(filename)

        if person_name not in grouped_data:
            grouped_data[person_name] = []

        grouped_data[person_name].append({
            'path': path,
            'scenario': scenario,
            'id': f_id
        })

    total_people = len(grouped_data)
    print(f"✅ พบข้อมูลทั้งหมด {len(file_paths)} ไฟล์ แบ่งเป็น {total_people} คน: {list(grouped_data.keys())}")

    # --- Step 2: Process Each Person ---
    for i, (person, data_list) in enumerate(grouped_data.items(), 1):
        print(f"\n" + "=" * 60)
        print(f"   👤 Processing Person {i}/{total_people}: {person}")
        print("=" * 60)

        person_output_files = []

        # Loop ข้อมูลย่อยของคนๆ นั้น (10 ชุด)
        for idx, item in enumerate(data_list, 1):
            print(f"   [{idx}/{len(data_list)}] วิเคราะห์: {item['scenario']} | ID: {item['id']}")
            out_files = process_single_file(item['path'], item['scenario'], person, item['id'])
            person_output_files.extend(out_files)

        # --- Step 3: Ask for Destination (Per Person) ---
        if person_output_files:
            print(f"\n💾 เสร็จสิ้นข้อมูลของ '{person}'")
            print(f"   👉 กรุณาเลือกโฟลเดอร์เก็บไฟล์ของ {person} ที่หน้าต่าง Popup...")

            dest_dir = filedialog.askdirectory(title=f"เลือกที่เก็บผลลัพธ์ของ: {person}")

            if dest_dir:
                print(f"📦 กำลังย้าย {len(person_output_files)} ไฟล์ ไปยัง {dest_dir}...")
                for f in person_output_files:
                    try:
                        shutil.move(f, os.path.join(dest_dir, os.path.basename(f)))
                    except:
                        pass
                print(f"✅ เรียบร้อยสำหรับ {person}")
            else:
                print(f"⚠️ ไม่ได้เลือกที่เก็บ ไฟล์จะกองอยู่ที่นี่")

    print("\n🎉🎉🎉 เสร็จสิ้นกระบวนการทั้งหมด! 🎉🎉🎉")


# ==========================================
# Main
# ==========================================
if __name__ == "__main__":
    while True:
        print("\n" + "=" * 40)
        print("   SENSOR AUTO-ANALYZER (BATCH MODE)")
        print("=" * 40)
        print(" [1] 🏗️ Build Global Standard (ทำครั้งแรก)")
        print(" [2] 🚀 Run Auto Analysis (ใส่ทุกไฟล์ -> แยกคนเอง -> ถามที่เก็บทีละคน)")
        print(" [0] Exit")

        c = input("Select: ").strip()
        if c == '1':
            build_global_standard()
        elif c == '2':
            analyze_auto_mode()
        elif c == '0':
            break