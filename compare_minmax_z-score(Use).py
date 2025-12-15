import pandas as pd
import os
import shutil
import tkinter as tk
from tkinter import filedialog
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
import numpy as np
import time


# ==========================================
# 1. ตั้งค่าพื้นฐาน
# ==========================================
def set_thai_font():
    if os.name == 'nt':
        plt.rcParams['font.family'] = 'Tahoma'
    else:
        plt.rcParams['font.family'] = 'Ayuthaya'
    plt.rcParams['axes.unicode_minus'] = False


# ==========================================
# 2. Core Logic (Calculation)
# ==========================================

def calculate_svm(df, prefix):
    cols = [f'{prefix}_X', f'{prefix}_Y', f'{prefix}_Z']
    if all(col in df.columns for col in cols):
        svm_col = f'{prefix}_SVM'
        df[svm_col] = (df[cols[0]] ** 2 + df[cols[1]] ** 2 + df[cols[2]] ** 2) ** 0.5
        return svm_col
    return None


def calculate_jerk(df, svm_col, time_col):
    dt_series = pd.Series(time_col).diff()
    avg_dt = dt_series.mean() if dt_series.mean() > 0 else 0.02

    # Smooth ลด Noise
    svm_smooth = df[svm_col].rolling(window=5, center=True).mean().bfill().ffill()

    jerk_col = f'{svm_col}_Jerk'
    df[jerk_col] = (svm_smooth.diff() / avg_dt).abs().fillna(0)
    return jerk_col


def calculate_hybrid(df, svm_col, jerk_col):
    hybrid_col = f'{svm_col}_Hybrid'
    df[hybrid_col] = df[svm_col] * df[jerk_col]
    return hybrid_col


def parse_time_column(df):
    if 'Time' not in df.columns: return df.index.to_numpy()
    try:
        time_str = df['Time'].astype(str)

        def t2s(t):
            p = t.strip().split(':')
            if len(p) == 2:
                return float(p[0]) * 60 + float(p[1])
            elif len(p) == 3:
                return float(p[0]) * 3600 + float(p[1]) * 60 + float(p[2])
            return 0.0

        sec = time_str.apply(t2s).to_numpy()
        return sec - sec[0] if len(sec) > 0 else df.index.to_numpy()
    except:
        return df.index.to_numpy()


# ==========================================
# 3. Scaling Functions
# ==========================================

def scale_minmax(series):
    """Min-Max (0-1)"""
    if series.max() == series.min(): return series.apply(lambda x: 0.0)
    return (series - series.min()) / (series.max() - series.min())


def scale_zscore(series):
    """Z-Score (Mean=0, SD=1)"""
    std = series.std()
    if std == 0: return series.apply(lambda x: 0.0)
    return (series - series.mean()) / std


# ==========================================
# 4. Visualization (แยก 3 ไฟล์ ไฟล์ละ 6 กราฟ)
# ==========================================

def plot_metric_comparison(df, filename_base, header_title, time_axis, metric_type):
    """
    ฟังก์ชันอเนกประสงค์สำหรับพล็อตกราฟเปรียบเทียบ Scaling
    metric_type: 'SVM', 'Jerk', หรือ 'Hybrid'
    """
    try:
        set_thai_font()
        SMOOTH_WINDOW = 7

        # ตั้งชื่อรูปและเลือกคอลัมน์ตามประเภทที่ส่งมา
        if metric_type == 'SVM':
            suffix = '_SVM'
            desc = 'Magnitude (แรงรวม)'
        elif metric_type == 'Jerk':
            suffix = '_SVM_Jerk'
            desc = 'Jerk (ความกระตุก)'
        elif metric_type == 'Hybrid':
            suffix = '_SVM_Hybrid'
            desc = 'Hybrid (Impact Index)'
        else:
            return None

        fig, axs = plt.subplots(3, 2, figsize=(16, 12))
        fig.suptitle(f'{metric_type} Analysis: {header_title}', fontsize=16, fontweight='bold')

        axs[0, 0].set_title(f"Min-Max (0-1)", fontsize=14, color='blue', fontweight='bold')
        axs[0, 1].set_title(f"Z-Score (SD)", fontsize=14, color='green', fontweight='bold')

        sensors = ['ACC', 'GYRO', 'MAG']

        def smooth(s):
            return s.rolling(window=SMOOTH_WINDOW, center=True, min_periods=1).mean()

        for i, sensor in enumerate(sensors):
            col_name = f'{sensor}{suffix}'  # เช่น ACC_SVM หรือ ACC_SVM_Jerk

            if col_name in df.columns:
                data = df[col_name]

                # --- Left: Min-Max ---
                mm_data = scale_minmax(data)
                axs[i, 0].plot(time_axis, smooth(mm_data), color='tab:blue', linewidth=1.2)
                axs[i, 0].set_ylabel(f'{sensor} {metric_type}\n(0-1)')
                axs[i, 0].grid(True, linestyle='--', alpha=0.5)
                axs[i, 0].set_ylim(-0.05, 1.05)

                # --- Right: Z-Score ---
                zs_data = scale_zscore(data)
                axs[i, 1].plot(time_axis, smooth(zs_data), color='tab:green', linewidth=1.2)
                axs[i, 1].set_ylabel(f'{sensor} {metric_type}\n(SD)')
                axs[i, 1].grid(True, linestyle='--', alpha=0.5)
                axs[i, 1].axhline(0, color='black', lw=0.8, alpha=0.3)  # เส้นศูนย์กลาง
            else:
                axs[i, 0].text(0.5, 0.5, 'No Data', ha='center');
                axs[i, 1].text(0.5, 0.5, 'No Data', ha='center')

        for ax in axs.flat: ax.set_xlabel('Time (s)')

        plt.tight_layout(rect=[0, 0.03, 1, 0.95])

        # บันทึกไฟล์แยกตามประเภท
        img_name = f"{filename_base}_{metric_type}_Compare.png"
        plt.savefig(img_name)
        plt.close()
        return img_name

    except Exception as e:
        print(f"Graph Error ({metric_type}): {e}")
        return None


# ==========================================
# 5. File Management
# ==========================================

def get_unique_filename(path):
    if not os.path.exists(path): return path
    base, ext = os.path.splitext(path)
    c = 1
    while True:
        new = f"{base} ({c}){ext}"
        if not os.path.exists(new): return new
        c += 1


def format_excel(ws, min_r, max_r, min_c, max_c):
    border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    for r in ws.iter_rows(min_r, max_r, min_c, max_c):
        for c in r:
            c.border = border
            c.alignment = Alignment('center', 'center')
            if isinstance(c.value, float): c.number_format = '0.0000'
    ws.column_dimensions['A'].width = 25
    for i in range(min_c + 1, max_c + 1): ws.column_dimensions[get_column_letter(i)].width = 16
    fill = PatternFill('solid', fgColor="D9D9D9")
    for i in range(min_c, max_c + 1):
        c = ws.cell(min_r, i);
        c.font = Font(bold=True);
        c.fill = fill


def organize_files(files):
    if not files: return
    print("\n" + "=" * 50 + "\n   🧹 Organize Output\n" + "=" * 50)
    excels = [f for f, t in files if t == 'excel']
    plots = [f for f, t in files if t == 'plot']

    def move_to_dir(file_list, type_name):
        if not file_list: return
        print(f"\n📂 เก็บไฟล์ {type_name}")
        root = tk.Tk();
        root.withdraw();
        root.attributes('-topmost', True)
        dest = filedialog.askdirectory(title=f"เลือกโฟลเดอร์เก็บ {type_name}")
        if dest:
            for f in file_list:
                try:
                    shutil.move(f, os.path.join(dest, os.path.basename(f)))
                except:
                    pass
            print(f"   ✔ ย้าย {len(file_list)} ไฟล์เรียบร้อย")

    move_to_dir(excels, "Excel Reports")
    move_to_dir(plots, "Graph Images")


# ==========================================
# 6. Main Logic
# ==========================================

def process_file(path, title, idx, total):
    print(f"\n[{idx}/{total}] ⚙️ Processing: {os.path.basename(path)}")
    try:
        df = pd.read_csv(path)
        t_sec = parse_time_column(df)

        # 1. Calculate All Features
        targets = []
        for s in ['ACC', 'GYRO', 'MAG']:
            svm = calculate_svm(df, s)
            if svm:
                targets.append(svm)
                jerk = calculate_jerk(df, svm, t_sec)
                targets.append(jerk)
                hybrid = calculate_hybrid(df, svm, jerk)
                targets.append(hybrid)

        # 2. Statistics
        stats = {}
        for c in targets:
            s = df[c]
            stats[c] = {'Mean': s.mean(), 'Max': s.max(), 'Min': s.min(), 'SD': s.std(), 'Kurt': s.kurt()}

        # 3. Save Excel
        xls = get_unique_filename(os.path.splitext(path)[0] + '_summary.xlsx')
        with pd.ExcelWriter(xls, engine='openpyxl') as w:
            df.to_excel(w, sheet_name='Raw Data', index=False)
            sdf = pd.DataFrame(stats)
            sdf.to_excel(w, sheet_name='Stats', startrow=2)
            ws = w.sheets['Stats']
            ws['A1'] = "Experiment:";
            ws['B1'] = title
            format_excel(ws, 3, 3 + len(stats), 1, sdf.shape[1] + 1)

        # 4. Generate 3 Separate Plot Files (Total 18 Graphs)
        generated_files = [(xls, 'excel')]
        base = os.path.splitext(xls)[0]

        # Plot 1: SVM Comparison
        img1 = plot_metric_comparison(df, base, title, t_sec, 'SVM')
        if img1: generated_files.append((img1, 'plot'))

        # Plot 2: Jerk Comparison
        img2 = plot_metric_comparison(df, base, title, t_sec, 'Jerk')
        if img2: generated_files.append((img2, 'plot'))

        # Plot 3: Hybrid Comparison
        img3 = plot_metric_comparison(df, base, title, t_sec, 'Hybrid')
        if img3: generated_files.append((img3, 'plot'))

        return generated_files

    except Exception as e:
        print(f"Error: {e}");
        return []


def analyze_mode():
    scenarios = ["ปกติ", "ช้า", "เอียงขวา", "เอียงซ้าย", "ซ้ายขวา", "หยุดชะงัก"]
    print("\n--- เลือกสถานการณ์ ---")
    for i, s in enumerate(scenarios, 1): print(f" [{i}] {s}")

    while True:
        c = input("เลือก: ").strip()
        if c.isdigit() and 1 <= int(c) <= 6:
            title = f"{scenarios[int(c) - 1]} : {input('ชื่อ: ').strip()}"
            break

    root = tk.Tk();
    root.withdraw();
    root.attributes('-topmost', True)
    paths = filedialog.askopenfilenames(title="เลือกไฟล์ CSV", filetypes=[("CSV", "*.csv")])
    if paths:
        outs = []
        for i, p in enumerate(paths, 1): outs.extend(process_file(p, title, i, len(paths)))
        organize_files(outs)


def batch_rename():
    root = tk.Tk();
    root.withdraw();
    root.attributes('-topmost', True)
    paths = filedialog.askopenfilenames(title="เลือกไฟล์เปลี่ยนชื่อ")
    if not paths: return
    files = sorted([(p, os.path.getmtime(p)) for p in paths], key=lambda x: x[1])
    base = input("ชื่อไฟล์ใหม่: ").strip() or "Data"
    for i, (old, _) in enumerate(files, 1):
        new = os.path.join(os.path.dirname(old), f"{base}_{i:02d}.csv")
        try:
            os.rename(old, new); print(f"Renamed: {os.path.basename(new)}")
        except:
            pass


if __name__ == "__main__":
    while True:
        print("\n=== Full Sensor Analysis (18 Graphs Edition) ===")
        print(" [1] Batch Rename")
        print(" [2] Analyze (Generate SVM, Jerk, Hybrid plots)")
        print(" [0] Exit")
        c = input("Select: ").strip()
        if c == '1':
            batch_rename()
        elif c == '2':
            analyze_mode()
        elif c == '0':
            break
