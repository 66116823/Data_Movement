import pandas as pd
import os
import shutil
import tkinter as tk
from tkinter import filedialog
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
import numpy as np
import time


# ==========================================
# ตั้งค่าฟอนต์ภาษาไทยสำหรับกราฟ
# ==========================================
def set_thai_font():
    if os.name == 'nt':  # Windows
        plt.rcParams['font.family'] = 'Tahoma'
    else:  # Mac / Linux
        plt.rcParams['font.family'] = 'Ayuthaya'
    plt.rcParams['axes.unicode_minus'] = False


# ==========================================
# ส่วนที่ 1: ฟังก์ชันคำนวณและ Normalization
# ==========================================

def calculate_svm(df, prefix):
    cols = [f'{prefix}_X', f'{prefix}_Y', f'{prefix}_Z']
    if all(col in df.columns for col in cols):
        svm_col = f'{prefix}_SVM'
        df[svm_col] = (df[cols[0]] ** 2 + df[cols[1]] ** 2 + df[cols[2]] ** 2) ** 0.5
        return svm_col
    return None


def normalize_series(series):
    if series.max() == series.min():
        return series.apply(lambda x: 0.0)
    return (series - series.min()) / (series.max() - series.min())


def parse_time_column(df):
    if 'Time' not in df.columns: return df.index.to_numpy()
    try:
        time_str = df['Time'].astype(str)

        def time_to_seconds(t_str):
            parts = t_str.strip().split(':')
            if len(parts) == 2:
                return float(parts[0]) * 60 + float(parts[1])
            elif len(parts) == 3:
                return float(parts[0]) * 3600 + float(parts[1]) * 60 + float(parts[2])
            else:
                return 0.0

        seconds_array = time_str.apply(time_to_seconds).to_numpy()
        if len(seconds_array) > 0:
            return seconds_array - seconds_array[0]
        else:
            return df.index.to_numpy()
    except Exception:
        return df.index.to_numpy()


# ==========================================
# ส่วนที่ 2: ฟังก์ชันจัดการไฟล์และ Excel
# ==========================================

def get_unique_filename(filepath):
    if not os.path.exists(filepath): return filepath
    base, ext = os.path.splitext(filepath)
    counter = 1
    while True:
        new_filepath = f"{base} ({counter}){ext}"
        if not os.path.exists(new_filepath): return new_filepath
        counter += 1


def format_excel_table(worksheet, min_row, max_row, min_col, max_col):
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    for row in worksheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if isinstance(cell.value, (int, float)): cell.number_format = '0.0000'
    worksheet.column_dimensions['A'].width = 20
    for col_idx in range(min_col + 1, max_col + 1):
        worksheet.column_dimensions[get_column_letter(col_idx)].width = 16
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    for col_idx in range(min_col, max_col + 1):
        cell = worksheet.cell(row=min_row, column=col_idx)
        cell.font = Font(bold=True);
        cell.fill = header_fill


# ==========================================
# ส่วนที่ 3: ฟังก์ชัน Plot กราฟ (Normalized)
# ==========================================

def plot_normalized_graph(df, filename_base, header_title, time_axis):
    try:
        set_thai_font()
        SMOOTH_WINDOW = 7

        fig, axs = plt.subplots(3, 1, figsize=(12, 12))
        fig.suptitle(f'Normalized SVM Analysis (0-1 Scale): {header_title}', fontsize=16, fontweight='bold')
        locator = ticker.MaxNLocator(integer=True)

        def smooth_data(series, window):
            return series.rolling(window=window, center=True, min_periods=1).mean()

        def plot_subplot(ax, data_col, color, label_text, title_text):
            if data_col in df.columns:
                data_to_plot = df[data_col]
                smoothed = smooth_data(data_to_plot, SMOOTH_WINDOW)
                ax.plot(time_axis, smoothed, color=color, label=f'Norm SVM (Smooth={SMOOTH_WINDOW})', linewidth=1.5)
                ax.set_ylabel('Norm. Mag (0-1)')
                ax.set_title(title_text, fontsize=12)
                ax.set_xlabel('Time (seconds)')
                ax.set_ylim(-0.05, 1.05)
                ax.xaxis.set_major_locator(locator)
                ax.grid(True, linestyle='--', alpha=0.6)
                ax.legend(loc='upper right')

        plot_subplot(axs[0], 'ACC_SVM_Norm', 'blue', 'ACC', 'Accelerometer: Normalized Pattern (รูปแบบความเร่ง)')
        plot_subplot(axs[1], 'GYRO_SVM_Norm', 'green', 'GYRO', 'Gyroscope: Normalized Pattern (รูปแบบการหมุน)')
        plot_subplot(axs[2], 'MAG_SVM_Norm', 'red', 'MAG', 'Magnetometer: Normalized Pattern (รูปแบบแม่เหล็ก)')

        plt.tight_layout(rect=[0, 0.03, 1, 0.96], h_pad=3.0)
        image_filename = filename_base + '_norm_plot.png'
        plt.savefig(image_filename)
        plt.close()
        return image_filename
    except Exception as e:
        print(f"   ⚠️ ไม่สามารถสร้างกราฟได้: {e}")
        return None


# ==========================================
# ส่วนที่ 4: ฟังก์ชันจัดระเบียบไฟล์ (New Feature)
# ==========================================

def select_directory_dialog(prompt_title):
    """เปิดหน้าต่างเลือก Folder และคืนค่า Path"""
    print(f"\n📂 {prompt_title}")
    print("   (หน้าต่างเลือกโฟลเดอร์กำลังเปิดอยู่...)")
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    folder_selected = filedialog.askdirectory(title=prompt_title)
    if folder_selected:
        print(f"   ✅ เลือก: {folder_selected}")
    else:
        print("   ❌ ไม่ได้เลือกโฟลเดอร์ (ข้ามการย้าย)")
    return folder_selected


def organize_output_files(source_dir, generated_files):
    """
    ย้ายไฟล์ผลลัพธ์ไปยังโฟลเดอร์ปลายทาง
    generated_files: list ของ tuple (file_path, file_type)
    file_type: 'excel' หรือ 'plot'
    """
    if not generated_files:
        return

    print("\n" + "=" * 60)
    print("   🧹 ขั้นตอนการจัดเก็บไฟล์ (Organize Output)")
    print("=" * 60)

    # 1. จัดการไฟล์ Excel (_summary.xlsx)
    excel_files = [f for f, t in generated_files if t == 'excel']
    if excel_files:
        print(f"\n1️⃣  พบไฟล์ Excel สรุปผลจำนวน {len(excel_files)} ไฟล์")
        dest_excel = select_directory_dialog("เลือกโฟลเดอร์เก็บไฟล์ Excel (_summary.xlsx)")

        if dest_excel:
            print("   กำลังย้ายไฟล์ Excel...")
            for f_path in excel_files:
                try:
                    shutil.move(f_path, os.path.join(dest_excel, os.path.basename(f_path)))
                    print(f"    - ย้าย: {os.path.basename(f_path)}")
                except Exception as e:
                    print(f"    ❌ ย้ายไม่ได้: {os.path.basename(f_path)} ({e})")

    # 2. จัดการไฟล์รูปภาพ (_norm_plot.png)
    plot_files = [f for f, t in generated_files if t == 'plot']
    if plot_files:
        print(f"\n2️⃣  พบไฟล์กราฟรูปภาพจำนวน {len(plot_files)} ไฟล์")
        dest_plot = select_directory_dialog("เลือกโฟลเดอร์เก็บไฟล์รูปภาพ (_norm_plot.png)")

        if dest_plot:
            print("   กำลังย้ายไฟล์รูปภาพ...")
            for f_path in plot_files:
                try:
                    shutil.move(f_path, os.path.join(dest_plot, os.path.basename(f_path)))
                    print(f"    - ย้าย: {os.path.basename(f_path)}")
                except Exception as e:
                    print(f"    ❌ ย้ายไม่ได้: {os.path.basename(f_path)} ({e})")

    print("\n✅ เสร็จสิ้นขั้นตอนการจัดเก็บไฟล์")


# ==========================================
# ส่วนที่ 5: Logic หลัก
# ==========================================

def batch_rename_mode():
    print("\n" + "=" * 60);
    print("   📂 โหมดเปลี่ยนชื่อไฟล์อัตโนมัติ");
    print("=" * 60)
    root = tk.Tk();
    root.withdraw();
    root.attributes('-topmost', True)
    file_paths = filedialog.askopenfilenames(title="เลือกไฟล์ CSV", filetypes=[("CSV Files", "*.csv")])
    if not file_paths: return

    files_with_time = [(f, os.path.getmtime(f)) for f in file_paths]
    files_with_time.sort(key=lambda x: x[1])

    base_name = input("ตั้งชื่อกลุ่มไฟล์ใหม่ (เช่น Patient_A): ").strip() or "Data"
    count = 0
    for i, (old_path, timestamp) in enumerate(files_with_time, 1):
        directory = os.path.dirname(old_path)
        new_filename = f"{base_name}_{i:02d}{os.path.splitext(old_path)[1]}"
        new_path = os.path.join(directory, new_filename)
        try:
            if old_path != new_path:
                os.rename(old_path, new_path)
                print(f"   ✔ {os.path.basename(old_path)} -> {new_filename}")
                count += 1
        except Exception as e:
            print(f"Error: {e}")
    print(f"🎉 เสร็จสิ้น {count} ไฟล์!");
    input("กด Enter กลับเมนู...")


def get_experiment_info():
    scenarios = ["กรณีเดินปกติ", "กรณีเดินปกติช้า", "กรณีเดินเอียงขวา", "กรณีเดินเอียงซ้าย", "กรณีเดินเอียงซ้ายขวา",
                 "กรณีเดินหยุดชะงัก"]
    print("\n--- เลือกสถานการณ์ ---")
    for i, s in enumerate(scenarios, 1): print(f" [{i}] {s}")
    while True:
        c = input("เลือก (1-6): ").strip()
        if c.isdigit() and 1 <= int(c) <= 6: return f"{scenarios[int(c) - 1]} : {input('ชื่อผู้ทดลอง: ').strip()}"


def process_sensor_file(file_path, header_title, current_idx, total_files):
    filename = os.path.basename(file_path)
    print(f"\n[{current_idx}/{total_files}] 📄 กำลังประมวลผล: {filename}")

    generated_output = []  # เก็บรายชื่อไฟล์ที่สร้างขึ้นเพื่อส่งไปย้ายทีหลัง

    try:
        df = pd.read_csv(file_path)
        elapsed_time = parse_time_column(df)

        svm_cols = []
        for sensor in ['ACC', 'GYRO', 'MAG']:
            raw_col = calculate_svm(df, sensor)
            if raw_col:
                svm_cols.append(raw_col)
                norm_col = f'{raw_col}_Norm'
                df[norm_col] = normalize_series(df[raw_col])

        target_stats_cols = [
            'ACC_X', 'ACC_Y', 'ACC_Z', 'ACC_SVM',
            'GYRO_X', 'GYRO_Y', 'GYRO_Z', 'GYRO_SVM',
            'MAG_X', 'MAG_Y', 'MAG_Z', 'MAG_SVM'
        ]
        valid_stats_cols = [c for c in target_stats_cols if c in df.columns]

        if not valid_stats_cols:
            print("   ❌ ไม่พบคอลัมน์ข้อมูล")
            return []

        stats_data = {}
        for col in valid_stats_cols:
            series = pd.to_numeric(df[col], errors='coerce')
            stats_data[col] = {
                'Mean': series.mean(), 'Median': series.median(), 'SD': series.std(),
                'Skewness': series.skew(), 'Kurtosis': series.kurt(),
                'Max': series.max(), 'Min': series.min()
            }
        summary_df = pd.DataFrame(stats_data)

        base_filename_full = os.path.splitext(file_path)[0] + '_summary.xlsx'
        final_excel_filename = get_unique_filename(base_filename_full)
        final_base_name = os.path.splitext(final_excel_filename)[0]

        start_row = 3
        with pd.ExcelWriter(final_excel_filename, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Raw Data', index=False)
            summary_df.to_excel(writer, sheet_name='Summary Stats', startrow=start_row - 1)
            ws = writer.sheets['Summary Stats']
            ws['A1'] = "การทดลอง:";
            ws['B1'] = header_title
            ws['A1'].font = Font(bold=True, size=12);
            ws['B1'].font = Font(bold=True, size=12)
            ws['D1'] = "Graph Type: Normalized (0-1)";
            ws['D1'].font = Font(italic=True, color="555555")
            max_r = start_row + summary_df.shape[0];
            max_c = summary_df.shape[1] + 1
            format_excel_table(ws, min_row=start_row, max_row=max_r, min_col=1, max_col=max_c)

        print(f"   ✔ Excel Created: {os.path.basename(final_excel_filename)}")
        generated_output.append((final_excel_filename, 'excel'))  # เก็บเข้า List

        image_file = plot_normalized_graph(df, final_base_name, header_title, elapsed_time)
        if image_file:
            print(f"   ✔ Graph Created: {os.path.basename(image_file)}")
            generated_output.append((image_file, 'plot'))  # เก็บเข้า List

    except Exception as e:
        print(f"   ❌ Error: {e}")

    return generated_output


def analyze_data_mode():
    title = get_experiment_info()
    root = tk.Tk();
    root.withdraw();
    root.attributes('-topmost', True)

    print("\n⏳ กรุณาเลือกไฟล์ CSV ที่ต้องการวิเคราะห์ (เลือกได้หลายไฟล์)...")
    paths = filedialog.askopenfilenames(title="เลือกไฟล์ CSV", filetypes=[("CSV Files", "*.csv")])

    if paths:
        print(f"\n📦 เลือก {len(paths)} ไฟล์")

        all_generated_files = []  # เก็บรวมทุกไฟล์ที่สร้างจากทุก loop

        # 1. Loop ประมวลผล
        for i, p in enumerate(paths, 1):
            output_files = process_sensor_file(p, title, i, len(paths))
            all_generated_files.extend(output_files)

        print("\n" + "=" * 50)
        print("🎉 ประมวลผลข้อมูลเสร็จสมบูรณ์!")
        print("=" * 50)

        # 2. เข้าสู่โหมดย้ายไฟล์ (Organize)
        if all_generated_files:
            # หา source directory จากไฟล์แรกที่เลือก (สมมติว่าไฟล์ output อยู่ที่เดียวกับ input)
            source_dir = os.path.dirname(paths[0])
            organize_output_files(source_dir, all_generated_files)

        input("\nกด Enter เพื่อกลับสู่เมนูหลัก...")
    else:
        print("❌ ไม่ได้เลือกไฟล์")


if __name__ == "__main__":
    while True:
        print("\n=== Sensor Analysis (Normalized SVM) ===")
        print(" [1] เปลี่ยนชื่อไฟล์ (Batch Rename)")
        print(" [2] วิเคราะห์ข้อมูล (Normalize + Plot + Organize)")
        print(" [0] ออก")
        c = input("เลือก: ").strip()
        if c == '1':
            batch_rename_mode()
        elif c == '2':
            analyze_data_mode()
        elif c == '0':
            break
