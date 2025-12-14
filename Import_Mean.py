import pandas as pd
import sys
import os
import tkinter as tk
from tkinter import filedialog
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font


def find_header_row(file_path, sheet_name):
    """ฟังก์ชันช่วยหาว่า Header อยู่บรรทัดไหน โดยหาคำว่า 'ACC_X'"""
    try:
        df_preview = pd.read_excel(file_path, sheet_name=sheet_name, nrows=10, header=None)
        for idx, row in df_preview.iterrows():
            row_str = row.astype(str).str.upper().tolist()
            if 'ACC_X' in row_str:
                return idx
        return 0
    except Exception as e:
        return 0


def process_files(file_paths):
    target_columns = [
        'ACC_X', 'ACC_Y', 'ACC_Z',
        'GYRO_X', 'GYRO_Y', 'GYRO_Z',
        'MAG_X', 'MAG_Y', 'MAG_Z'
    ]

    data_rows = []
    print(f"กำลังประมวลผลทั้งหมด {len(file_paths)} ไฟล์...")

    # --- เริ่มวนลูปอ่านข้อมูลจากไฟล์ต้นฉบับ ---
    for index, file_path in enumerate(file_paths):
        try:
            # 1. หาตำแหน่ง Header และอ่านไฟล์
            header_row_idx = find_header_row(file_path, 'Summary Stats')
            df = pd.read_excel(file_path, sheet_name='Summary Stats', header=header_row_idx)
            df.columns = df.columns.str.strip().str.upper()

            # 2. หาแถวที่มีค่า Mean
            mean_row = None
            for idx, row in df.iterrows():
                first_few_cols = row.iloc[:3].astype(str).str.lower().tolist()
                if any('mean' in str(x) for x in first_few_cols):
                    mean_row = row
                    break

            # 3. ดึงข้อมูล
            if mean_row is not None:
                mean_values = {}
                for col in target_columns:
                    mean_values[col] = mean_row[col] if col in df.columns else None
            else:
                print(f"⚠️ คำเตือน: หาแถว 'Mean' ไม่เจอในไฟล์ {os.path.basename(file_path)}")
                mean_values = {col: None for col in target_columns}

            # เก็บข้อมูลลง List
            file_name = os.path.basename(file_path)
            row_data = [file_name]
            for col in target_columns:
                row_data.append(mean_values.get(col))

            data_rows.append(row_data)

        except Exception as e:
            print(f"❌ Error ไฟล์ {file_path}: {e}")
            row_data = [os.path.basename(file_path)] + [None] * 9
            data_rows.append(row_data)

    # --- ส่วนที่ 2: ให้ผู้ใช้เลือกที่จัดเก็บไฟล์ (Save As) ---
    print("ประมวลผลเสร็จสิ้น กรุณาเลือกที่บันทึกไฟล์...")

    # กำหนดให้เริ่มหาที่อยู่จากโฟลเดอร์ของไฟล์แรกที่นำเข้า (เพื่อความสะดวก)
    initial_dir = os.path.dirname(file_paths[0]) if file_paths else os.path.expanduser("~")

    full_save_path = filedialog.asksaveasfilename(
        title="บันทึกไฟล์สรุปผล (Save As)",
        initialdir=initial_dir,
        initialfile="Final_Summary_Result.xlsx",
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")]
    )

    # ตรวจสอบว่าผู้ใช้กดยกเลิกหรือไม่ (ถ้าไม่เลือก path จะเป็นค่าว่าง)
    if not full_save_path:
        print("❌ ยกเลิกการบันทึกไฟล์")
        return

    # --- ส่วนสร้างไฟล์ Excel ผลลัพธ์ ---
    header_columns = ["ชื่อไฟล์"] + target_columns
    result_df = pd.DataFrame(data_rows, columns=header_columns)

    try:
        with pd.ExcelWriter(full_save_path, engine='openpyxl') as writer:
            result_df.to_excel(writer, sheet_name='Result', index=False, startrow=1)

            workbook = writer.book
            worksheet = writer.sheets['Result']

            # จัดรูปแบบ (Formatting)
            worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(header_columns))
            cell = worksheet.cell(row=1, column=1)
            cell.value = "สรุปผลค่าเฉลี่ย (Mean Summary)"
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(bold=True, size=14)
            cell.fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")

            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                 bottom=Side(style='thin'))

            for row in range(1, len(data_rows) + 3):
                for col in range(1, len(header_columns) + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.border = thin_border
                    if row > 2 and col > 1:
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                        cell.number_format = '0.0000'
                    elif row <= 2:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.font = Font(bold=True)

        # --- ส่วนแสดงผลและเปิดโฟลเดอร์ ---
        output_folder = os.path.dirname(full_save_path)
        print("\n" + "=" * 60)
        print(f"✅ สำเร็จ! ไฟล์ถูกบันทึกเรียบร้อยแล้ว")
        print(f"📂 ตำแหน่ง: {full_save_path}")
        print("=" * 60 + "\n")

        # เปิดโฟลเดอร์ให้อัตโนมัติ (Windows)
        if os.name == 'nt':
            os.startfile(output_folder)

    except PermissionError:
        print(f"\n❌ Error: ไม่สามารถบันทึกไฟล์ได้! กรุณาปิดไฟล์ปลายทางก่อนรันโปรแกรมใหม่")
    except Exception as e:
        print(f"\n❌ Error ระหว่างบันทึกไฟล์: {e}")


def select_files_and_run():
    root = tk.Tk()
    root.withdraw()  # ซ่อนหน้าต่างหลัก

    # ทำให้หน้าต่าง Dialog เด้งขึ้นมาอยู่ข้างหน้าสุดเสมอ
    root.attributes('-topmost', True)

    print("กรุณาเลือกไฟล์ Excel จากหน้าต่างที่ปรากฏขึ้น...")
    file_paths = filedialog.askopenfilenames(
        title="เลือกไฟล์ Excel ที่ต้องการสรุปข้อมูล",
        filetypes=[("Excel files", "*.xlsx *.xls")]
    )
    if file_paths:
        process_files(list(file_paths))
    else:
        print("ยกเลิกการทำงาน: ไม่ได้เลือกไฟล์ใดๆ")

if __name__ == "__main__":
    select_files_and_run()
